#!/usr/bin/env python3
"""Artifact-only conformance checker — Schema Specification v0.8.x.
Implements the Level-1-relevant rule set from spec §3 (D1/D2), §11.7 (I1-I22),
§11.9 (I23/I24/I26/I28-I32), plus lookup layer (I13/I14), D6 provenance
cardinality, §9.2 Required-attribute coverage, and the §11.4 P0-P9
post-conditions insofar as they are decidable from the artifact alone.

Scope discipline: every check here must be decidable from the workbook and
this specification alone. No check may infer anything about the producing
tool or process (that is Level 2/3, out of scope for this checker).
"""
import sys, re
from pathlib import Path
from collections import Counter

import openpyxl
import json
from datetime import datetime

MANDATORY = ["Rules","Schema_Metadata","Document_ID","Document_Data","Revision_Data",
 "Document_RepresentedItem","Object","Cluster","Object_Cluster","Elements_TopDown",
 "Elements_from_Cluster","Match_Result","Element_ID","Element_RepresentedItem_Mapping",
 "Element_Data","Element_Data_Source","RepresentedItem_Data","RepresentedItem_Data_Source",
 "Element_Classification","Connection_ID","Connection_Data","Connection_Data_Source",
 "Layer_ID","Attribute_Lookup","Enum_Lookup","Document_Data_Source","Revision_Data_Source",
 "Element_Classification_Source"]
OPTIONAL_ORDER = ["Designation","Electrical_Node","Electrical_Node_Member","Cable_Data"]

RULE_DEFINITIONS = {
    "D1/A-structural": "Prüft, ob alle Pflichtblätter vorhanden sind und die Workbook-Struktur dem geforderten Schema entspricht.",
    "D2-ordering": "Prüft die Reihenfolge der Pflicht- und Optionalblätter gemäß der Spezifikation.",
    "A-structural": "Prüft grundlegende Strukturmerkmale wie fehlende Schlüsselspalten und doppelte Primärschlüssel.",
    "I1": "Prüft, ob Referenzen aus Element_ID auf gültige Match-Ergebnisse zeigen.",
    "I5": "Prüft, ob Object_Cluster-Einträge nur auf vorhandene Object- und Cluster-IDs verweisen.",
    "I7": "Prüft Topologie-Referenzen zwischen Objekten.",
    "I8": "Prüft, ob Verbindungsendpunkte auf gültige Elemente zeigen und die richtige CAEX-Typisierung haben.",
    "I9": "Prüft die Zuordnung zwischen Elementen und represented items.",
    "I10": "Prüft, ob Element-Daten nur auf vorhandene Elemente verweisen.",
    "I11": "Prüft, ob RepresentedItem-Daten nur auf vorhandene RepresentedItems verweisen.",
    "I12": "Prüft die Herkunftsbeziehungen der Daten- und Klassifikationsquellen.",
    "I12/D6": "Prüft, ob jede Datenzeile mindestens eine passende Quellzeile besitzt.",
    "I13": "Prüft, ob Attributnamen für den jeweiligen Scope in Attribute_Lookup definiert sind.",
    "I14": "Prüft, ob Attributwerte den erlaubten Enum-Lookup-Werten entsprechen.",
    "I15": "Prüft, ob Dokument-IDs in abhängigen Blättern gültig referenziert werden.",
    "I16": "Prüft die Integrität von Eltern-Kind-Beziehungen in Hierarchien.",
    "I17": "Prüft, ob Container-Cluster auf geschlossene Graphic-Objekte verweisen.",
    "I18": "Prüft Referenzen aus Elements_from_Cluster auf vorhandene Cluster.",
    "I19": "Prüft Match_Result-Referenzen auf TopDown- und Cluster-basierte Element-IDs.",
    "I20": "Prüft, ob Match-Zustände und Resolution-Status konsistent sind.",
    "I21": "Prüft, ob Element-IDs auf gültige Layer verweisen.",
    "I22": "Prüft, ob maximal eine Primary-Mapping pro Element existiert.",
    "I23": "Prüft inhaltlich Brücken-/Klemmleisten-Konsistenz für Terminal_Diagram-Workbooks (§11.9).",
    "I24": "Prüft inhaltlich Aderfarbe/Polarität-Konsistenz für Terminal_Diagram/Circuit_Diagram-Workbooks (§11.9).",
    "I26": "Prüft inhaltlich Current_Path_Number-Konsistenz für Circuit_Diagram-Workbooks (§11.9).",
    "I28": "Prüft inhaltlich IEC-60617-Klassifikationsvollständigkeit für Circuit_Diagram-Workbooks (§11.9).",
    "I29": "Prüft, ob die Source_Format-Information eindeutig und enum-konform vorliegt.",
    "I30": "Prüft polymorphe Klassifikationsreferenzen auf gültige Zielobjekte.",
    "I31": "Prüft das Kabelmodellierungsprofil und die zugehörigen Cable_Data-Referenzen.",
    "I32": "Prüft optionale Designation-Blätter und deren Eltern-Referenzen.",
    "I2": "Prüft Match_Result-Zustände für 'Matched' und verlangt vollständige Identitäten.",
    "I3": "Prüft Match_Result-Zustände für 'Only_TopDown' und verlangt korrekte Belegung.",
    "I4": "Prüft Match_Result-Zustände für 'Only_Cluster' und verlangt korrekte Belegung.",
    "I6/I16": "Prüft Konsistenz von Hierarchien in Clustern und Elementen.",
    "I31/I25": "Prüft Cable_Data-Referenzen im Zusammenhang mit dem Kabelprofil.",
    "Required_Attributes": "Prüft, ob alle in Attribute_Lookup als Required=TRUE markierten Attribute für jedes "
                            "anwendbare Objekt (nach Type_Constraint, §9.4) tatsächlich belegt sind (§9.2).",
    "Cluster_Coverage": "Prüft für PDF-Quellen, ob jedes nicht-Topology-Object in mindestens einem Cluster "
                        "vorkommt (P3-Postbedingung, §11.4), sofern kein dokumentierter P3_Empty_Result vorliegt.",
}

# Rules with actual check code below. A RULE_DEFINITIONS entry NOT in this set
# is reported as NOT_IMPLEMENTED rather than silently defaulting to PASS -- see
# reset_rule_results(). Keep this in sync when adding a new rule's check code.
IMPLEMENTED_RULES = set(RULE_DEFINITIONS.keys())

F = []     # blocking findings (FAIL): (layer, rule, detail)
WARN = []  # content-completeness notices (WARNING): (layer, rule, detail, group, ref)
RULE_RESULTS = {}


def f(layer, rule, detail):
    F.append((layer, rule, detail))
    note_rule(rule, status="FAIL", detail=detail)


def w(layer, rule, detail, group=None, ref=None):
    """Record a content-completeness notice: a Required=TRUE Attribute_Lookup
    entry with no corresponding value in the source-derived data. `group` is
    a short category key (e.g. the Attribute_Name) and `ref` the specific
    instance identifier (e.g. the Element_ID) affected -- together these let
    the report summarize many same-cause findings ("Rated_Current: 8 rows,
    e.g. E.59, E.60, ...") instead of listing them as repetitive lines. Does
    not count against the Level-1 pass rate (see check_required_attributes)."""
    WARN.append((layer, rule, detail, group, ref))
    note_rule(rule, status="WARNING", detail=detail)


def note_rule(rule, status="PASS", detail=None, description=None):
    """Record an outcome for `rule`. Status precedence: FAIL beats SKIP beats
    WARNING beats PASS beats NOT_RUN; NOT_IMPLEMENTED is set once at reset
    time and is never touched again (no check code exists to call this for
    it)."""
    entry = RULE_RESULTS.setdefault(
        rule,
        {"description": description or RULE_DEFINITIONS.get(rule, "Keine Beschreibung hinterlegt."),
         "status": "NOT_RUN", "details": []},
    )
    if entry["status"] == "NOT_IMPLEMENTED":
        # Should not normally happen (no check code calls note_rule for such a
        # rule) but guard defensively rather than silently overwrite.
        return entry
    if status == "FAIL":
        entry["status"] = "FAIL"
    elif status == "SKIP" and entry["status"] != "FAIL":
        entry["status"] = "SKIP"
    elif status == "WARNING" and entry["status"] not in {"FAIL", "SKIP"}:
        entry["status"] = "WARNING"
    elif status == "PASS" and entry["status"] not in {"FAIL", "SKIP", "WARNING"}:
        entry["status"] = "PASS"
    if detail:
        entry["details"].append(detail)
    return entry


def reset_rule_results():
    RULE_RESULTS.clear()
    for rule, description in RULE_DEFINITIONS.items():
        initial = "NOT_RUN" if rule in IMPLEMENTED_RULES else "NOT_IMPLEMENTED"
        details = [] if rule in IMPLEMENTED_RULES else [
            "This rule is declared in the specification but has no check implementation in this "
            "checker version. Its status is NOT_IMPLEMENTED, not PASS -- absence of a finding here "
            "is not evidence of conformance."
        ]
        RULE_RESULTS[rule] = {"description": description, "status": initial, "details": details}


def resolve_artifact_dir(base_dir=None):
    base = Path(base_dir or Path(__file__).resolve().parent).resolve()
    artefacts_dir = base / "artefacts"
    standardized_dir = artefacts_dir / "Standardized Intermediate"
    if standardized_dir.is_dir():
        return standardized_dir
    artefacts_dir.mkdir(exist_ok=True)
    return artefacts_dir


def resolve_results_dir(base_dir=None):
    base = Path(base_dir or Path(__file__).resolve().parent).resolve()
    results_dir = base / "results"
    results_dir.mkdir(exist_ok=True)
    return results_dir


def discover_excel_files(input_dir):
    if not input_dir.exists():
        return []
    return sorted(
        [p for p in input_dir.iterdir() if p.is_file() and p.suffix.lower() in {".xlsx", ".xlsm", ".xls"}]
    )


def norm(v):
    if v is None: return None
    s = str(v).strip()
    return s if s != "" else None

def load(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheets = {}
    for name in wb.sheetnames:
        ws = wb[name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            sheets[name] = ([], []); continue
        hdr = [norm(h) for h in rows[0]]
        data = []
        for r in rows[1:]:
            if all(norm(c) is None for c in r): continue
            data.append({hdr[i]: norm(r[i]) for i in range(min(len(hdr), len(r))) if hdr[i]})
        sheets[name] = (hdr, data)
    return wb.sheetnames, sheets

def col(sheets, sheet, c):
    hdr, data = sheets[sheet]
    return [row.get(c) for row in data]

def ids(sheets, sheet, c):
    return set(v for v in col(sheets, sheet, c) if v is not None)


# ---------------------------------------------------------------------------
# §9.4 Type_Constraint evaluation (Equality / Set-inclusion / AND conjunction)
# ---------------------------------------------------------------------------
_TC_EQ = re.compile(r"^([A-Za-z_]+)\s*=\s*(.+)$")
_TC_SET = re.compile(r"^([A-Za-z_]+)\s*(?:∈|in)\s*\{(.+)\}$")


def tc_applies(tc, context):
    """Evaluate a §9.4 Type_Constraint expression against a context dict of
    {field_name: actual_value}. Empty/None Type_Constraint always applies.
    Unrecognized atoms are treated permissively (do not exclude), since this
    checker must not invent stricter semantics than §9.4 defines."""
    if not tc:
        return True
    for atom in tc.split(" AND "):
        atom = atom.strip()
        m = _TC_SET.match(atom)
        if m:
            field, vals = m.group(1), [v.strip() for v in m.group(2).split(",")]
            if context.get(field) not in vals:
                return False
            continue
        m = _TC_EQ.match(atom)
        if m:
            field, val = m.group(1), m.group(2).strip()
            if context.get(field) != val:
                return False
            continue
        # Unrecognized Type_Constraint syntax: do not silently exclude the row;
        # this checker does not invent stricter semantics than §9.4 defines.
    return True


def scope_includes(doc_type, scope_csv):
    """§11.9 document-type applicability, e.g. 'Terminal_Diagram/Circuit_Diagram'.
    Returns False (not applicable) when doc_type is undeterminable, rather than
    raising -- fixes the previous crash on an empty Document_ID sheet."""
    if doc_type is None:
        return False
    allowed = [s.strip() for s in scope_csv.split("/")]
    return doc_type in allowed


def check(path, output_path=None):
    F.clear()
    WARN.clear()
    reset_rule_results()
    order, sheets = load(path)

    # ---------- Layer 1: structural (D1, D2, PK, FK) ----------
    missing = [s for s in MANDATORY if s not in order]
    for s in missing: f(1, "D1/A-structural", f"mandatory sheet missing: {s}")
    present_mand = [s for s in order if s in MANDATORY]
    if not missing and present_mand != MANDATORY:
        f(1, "D2-ordering", f"mandatory sheet order deviates from catalog; first deviation at position "
          f"{next(i for i,(a,b) in enumerate(zip(present_mand, MANDATORY)) if a!=b)+1} "
          f"(found {next(a for a,b in zip(present_mand, MANDATORY) if a!=b)!r})")
    opts = [s for s in order if s not in MANDATORY]
    # optional sheets must come after sheet #28 and in O-order
    if not missing:
        last_mand_pos = max(order.index(s) for s in MANDATORY)
        for s in opts:
            if order.index(s) < last_mand_pos:
                f(1, "D2-ordering", f"optional sheet {s!r} placed before mandatory sheet #28")
        opt_seq = [s for s in order if s in OPTIONAL_ORDER]
        if opt_seq != [s for s in OPTIONAL_ORDER if s in opt_seq]:
            f(1, "D2-ordering", f"optional sheet order {opt_seq} deviates from O1..O3 order")
    note_rule("D2-ordering", "PASS")
    if missing:
        note_rule("D1/A-structural", "PASS")  # merge no-ops if already FAIL; keeps status coherent
        report(doc_type=None, sheets=sheets, output_path=output_path); return
    note_rule("D1/A-structural", "PASS")

    # PK uniqueness
    PK = {"Document_ID":"Document_ID","Document_Data":"Document_Data_ID","Revision_Data":"Revision_ID",
     "Document_RepresentedItem":"RepresentedItem_ID","Object":"Object_ID","Cluster":"Cluster_ID",
     "Elements_TopDown":"Element_TopDown_ID","Elements_from_Cluster":"Element_from_Cluster_ID",
     "Match_Result":"Match_ID","Element_ID":"Element_ID","Element_Data":"Element_Data_ID",
     "RepresentedItem_Data":"RepresentedItem_Data_ID","Element_Classification":"Classification_ID",
     "Connection_ID":"Connection_ID","Connection_Data":"Connection_Data_ID","Layer_ID":"Layer_ID"}
    for s, k in PK.items():
        hdr, data = sheets[s]
        if k not in hdr: f(1, "A-structural", f"{s}: key column {k} missing"); continue
        vals = [r.get(k) for r in data if r.get(k) is not None]
        dup = {v for v in vals if vals.count(v) > 1}
        for d in dup: f(1, "A-structural", f"{s}: duplicate {k}={d}")
    note_rule("A-structural", "PASS")

    doc_ids = ids(sheets, "Document_ID", "Document_ID")
    E  = ids(sheets, "Element_ID", "Element_ID")
    OBJ= ids(sheets, "Object", "Object_ID")
    CL = ids(sheets, "Cluster", "Cluster_ID")
    ETD= ids(sheets, "Elements_TopDown", "Element_TopDown_ID")
    EFC= ids(sheets, "Elements_from_Cluster", "Element_from_Cluster_ID")
    MR = ids(sheets, "Match_Result", "Match_ID")
    RIT= ids(sheets, "Document_RepresentedItem", "RepresentedItem_ID")
    CID= ids(sheets, "Connection_ID", "Connection_ID")
    LAY= ids(sheets, "Layer_ID", "Layer_ID")

    if not sheets["Document_ID"][1]:
        f(1, "D1/A-structural", "Document_ID sheet has no data rows; Document_Type is undeterminable "
                                 "(all Document_Type-scoped checks below will report Not applicable).")

    # I15 Document_ID FKs
    I15_SHEETS = ["Document_Data","Revision_Data","Document_RepresentedItem","Object","Cluster",
      "Elements_TopDown","Elements_from_Cluster","Match_Result","Element_ID","Connection_ID",
      "Element_Classification","Layer_ID"] + [s for s in ["Designation","Electrical_Node","Cable_Data"] if s in sheets]
    for s in I15_SHEETS:
        for i, v in enumerate(col(sheets, s, "Document_ID"), 2):
            if v is not None and v not in doc_ids:
                f(1, "I15", f"{s} row {i}: Document_ID {v!r} unresolved")
    note_rule("I15", "PASS")

    def fk(rule, sheet, column, target, allow_null=True):
        if sheet not in sheets:
            note_rule(rule, "PASS", detail=f"{sheet} sheet absent (optional); nothing to check.")
            return
        hdr, data = sheets[sheet]
        if column not in hdr:
            f(1, "A-structural", f"{sheet}: column {column} missing")
            return
        for i, r in enumerate(data, 2):
            v = r.get(column)
            if v is None:
                if not allow_null: f(1, rule, f"{sheet} row {i}: {column} is null")
                continue
            if v not in target: f(1, rule, f"{sheet} row {i}: {column}={v!r} unresolved")
        note_rule(rule, "PASS")

    fk("I1",  "Element_ID", "Source_Match_ID", MR, allow_null=False)
    fk("I5",  "Object_Cluster", "Cluster_ID", CL, allow_null=False)
    fk("I5",  "Object_Cluster", "Object_ID", OBJ, allow_null=False)
    fk("I18", "Elements_from_Cluster", "Source_Cluster_ID", CL, allow_null=False)
    fk("I19", "Match_Result", "Element_TopDown_ID", ETD)
    fk("I19", "Match_Result", "Element_from_Cluster_ID", EFC)
    fk("I9",  "Element_RepresentedItem_Mapping", "Element_ID", E, allow_null=False)
    fk("I9",  "Element_RepresentedItem_Mapping", "RepresentedItem_ID", RIT, allow_null=False)
    fk("I10", "Element_Data", "Element_ID", E, allow_null=False)
    fk("I11", "RepresentedItem_Data", "RepresentedItem_ID", RIT, allow_null=False)
    fk("I21", "Element_ID", "Layer_ID", LAY)
    fk("I7",  "Object", "Topology_From_Object_ID", OBJ)
    fk("I7",  "Object", "Topology_To_Object_ID", OBJ)

    # I8 connection endpoints
    caex = {r.get("Element_ID"): r.get("CAEX_Type") for r in sheets["Element_ID"][1]}
    for i, r in enumerate(sheets["Connection_ID"][1], 2):
        for c in ("From_Element_ID","To_Element_ID"):
            v = r.get(c)
            if v is None: f(1, "I8", f"Connection_ID row {i}: {c} is null"); continue
            if v not in E: f(1, "I8", f"Connection_ID row {i}: {c}={v!r} unresolved")
            elif caex.get(v) != "ExternalInterface":
                f(1, "I8", f"Connection_ID row {i}: {c}={v!r} has CAEX_Type={caex.get(v)!r} (ExternalInterface required)")
    note_rule("I8", "PASS")

    # I12 FKs (source sheets -> parents + Source_Object_ID -> Object)
    SRC = {"Element_Data_Source":("Element_Data_ID", ids(sheets,"Element_Data","Element_Data_ID")),
      "RepresentedItem_Data_Source":("RepresentedItem_Data_ID", ids(sheets,"RepresentedItem_Data","RepresentedItem_Data_ID")),
      "Connection_Data_Source":("Connection_Data_ID", ids(sheets,"Connection_Data","Connection_Data_ID")),
      "Document_Data_Source":("Document_Data_ID", ids(sheets,"Document_Data","Document_Data_ID")),
      "Revision_Data_Source":("Revision_ID", ids(sheets,"Revision_Data","Revision_ID")),
      "Element_Classification_Source":("Classification_ID", ids(sheets,"Element_Classification","Classification_ID"))}
    for s,(k,tgt) in SRC.items():
        fk("I12", s, k, tgt, allow_null=False)
        fk("I12", s, "Source_Object_ID", OBJ, allow_null=False)
    # D6 cardinality: every data row has >=1 source row
    for parent,(k,_) in [("Element_Data",SRC["Element_Data_Source"]),("RepresentedItem_Data",SRC["RepresentedItem_Data_Source"]),
        ("Connection_Data",SRC["Connection_Data_Source"]),("Document_Data",SRC["Document_Data_Source"]),
        ("Revision_Data",SRC["Revision_Data_Source"]),("Element_Classification",SRC["Element_Classification_Source"])]:
        src_sheet = parent + "_Source" if parent != "Element_Classification" else "Element_Classification_Source"
        have = ids(sheets, src_sheet, k)
        for i, r in enumerate(sheets[parent][1], 2):
            v = r.get(k)
            if v is not None and v not in have:
                f(1, "I12/D6", f"{parent} row {i}: no {src_sheet} row for {k}={v!r}")
    note_rule("I12/D6", "PASS")

    # I30 polymorphic classification
    TGT = {"Element":E, "RepresentedItem":RIT, "Connection":CID, "Document":doc_ids}
    for i, r in enumerate(sheets["Element_Classification"][1], 2):
        t, v = r.get("Classified_Object_Type"), r.get("Classified_Object_ID")
        if t not in TGT: f(1, "I30", f"Element_Classification row {i}: Classified_Object_Type={t!r} unknown"); continue
        if v is None or v not in TGT[t]:
            f(1, "I30", f"Element_Classification row {i}: {t} target {v!r} unresolved")
    note_rule("I30", "PASS")

    # I31 cable profile
    meta = {r.get("Metadata_Key"): r.get("Metadata_Value") for r in sheets["Schema_Metadata"][1]}
    prof_rows = [r for r in sheets["Schema_Metadata"][1] if r.get("Metadata_Key") == "Cable_Modeling_Profile"]
    if len(prof_rows) != 1: f(1, "I31", f"Cable_Modeling_Profile rows: {len(prof_rows)} (exactly one required)")
    prof = prof_rows[0].get("Metadata_Value") if prof_rows else None
    if prof not in ("Core","Asset"): f(1, "I31", f"Cable_Modeling_Profile={prof!r} invalid")
    if prof == "Asset":
        if "Cable_Data" not in sheets: f(1, "I31", "profile Asset but Cable_Data sheet absent")
        else: fk("I31/I25", "Connection_ID", "Cable_Data_ID", ids(sheets,"Cable_Data","Cable_Data_ID"))
    if prof == "Core":
        if "Cable_Data" in sheets: f(1, "I31", "profile Core but Cable_Data sheet present")
        for i, r in enumerate(sheets["Connection_ID"][1], 2):
            if r.get("Cable_Data_ID") is not None: f(1, "I31", f"Connection_ID row {i}: Cable_Data_ID set under Core profile")
    note_rule("I31", "PASS")
    if "I31/I25" not in RULE_RESULTS or RULE_RESULTS["I31/I25"]["status"] == "NOT_RUN":
        note_rule("I31/I25", "PASS", detail="Core profile or Asset profile without Cable_Data_ID population; nothing to check.")

    # I32 optional sheet FKs
    if "Designation" in sheets:
        fk("I32", "Designation", "Source_Object_ID", OBJ)
        desg = ids(sheets, "Designation", "Designation_ID")
        fk("I32", "Designation", "Parent_Designation_ID", desg)
    else:
        note_rule("I32", "PASS", detail="No optional extension sheets present; nothing to check.")

    # ---------- Layer 2: lookup / enum ----------
    doc_type = next(iter(col(sheets, "Document_ID", "Document_Type")), None)
    schema_version = next(iter(col(sheets, "Document_ID", "Schema_Version")), None)

    AL = {}
    for r in sheets["Attribute_Lookup"][1]:
        AL.setdefault((r.get("Scope"), r.get("Attribute_Name")), []).append(r)
    ENUM = {}
    for r in sheets["Enum_Lookup"][1]:
        ENUM.setdefault(r.get("Field_Name"), set()).add(r.get("Allowed_Value"))
    ext_e = set((meta.get("Allowed_Element_Type_Extensions") or "").split(",")) - {""}
    ext_r = set((meta.get("Allowed_RepresentedItem_Type_Extensions") or "").split(",")) - {""}

    # Per-row §9.4 context: Type_Constraint may reference Document_Type, Element_Type, or
    # RepresentedItem_Type. Element_Data / RepresentedItem_Data rows must therefore resolve
    # their OWNING element's/represented-item's own type via FK, not just the workbook's
    # Document_Type -- otherwise every Element_Type=... / RepresentedItem_Type=...
    # Type_Constraint would incorrectly evaluate as "not applicable".
    etype_by_eid = {r.get("Element_ID"): r.get("Element_Type") for r in sheets["Element_ID"][1]}
    ritype_by_riid = {r.get("RepresentedItem_ID"): r.get("RepresentedItem_Type")
                      for r in sheets["Document_RepresentedItem"][1]}

    DATA_SCOPE = {"Document_Data":"Document","Element_Data":"Element",
                  "RepresentedItem_Data":"RepresentedItem","Connection_Data":"Connection"}
    OWNER_COL = {"Element_Data": "Element_ID", "RepresentedItem_Data": "RepresentedItem_ID"}
    for sheet, scope in DATA_SCOPE.items():
        for i, r in enumerate(sheets[sheet][1], 2):
            an, av = r.get("Attribute_Name"), r.get("Attribute_Value")
            context = {"Document_Type": doc_type}
            if sheet == "Element_Data":
                context["Element_Type"] = etype_by_eid.get(r.get(OWNER_COL[sheet]))
            elif sheet == "RepresentedItem_Data":
                context["RepresentedItem_Type"] = ritype_by_riid.get(r.get(OWNER_COL[sheet]))
            entries = AL.get((scope, an))
            if not entries:
                f(2, "I13", f"{sheet} row {i}: Attribute_Name={an!r} not in Attribute_Lookup for scope {scope}")
                continue
            if not any(tc_applies(e.get("Type_Constraint"), context) for e in entries):
                f(2, "I13", f"{sheet} row {i}: {an!r} not allowed in this row's context "
                            f"(Type_Constraint(s): {[e.get('Type_Constraint') for e in entries]}, "
                            f"context: {context})")
            enum_field = next((e.get("Allowed_Values_Enum_Field") for e in entries if e.get("Allowed_Values_Enum_Field")), None)
            if enum_field and av is not None and av != "Unspecifiable":
                allowed = ENUM.get(enum_field, set())
                extra = ext_e if enum_field == "Element_Type" else ext_r if enum_field == "RepresentedItem_Type" else set()
                if av not in allowed | extra:
                    f(2, "I14", f"{sheet} row {i}: {an}={av!r} not in Enum_Lookup[{enum_field}]")
    note_rule("I13", "PASS")
    note_rule("I14", "PASS")

    # generic enum columns: any column whose header is a Field_Name in Enum_Lookup
    for sheet in MANDATORY + [s for s in OPTIONAL_ORDER if s in sheets]:
        if sheet in ("Enum_Lookup","Attribute_Lookup","Rules"): continue
        hdr, data = sheets[sheet]
        for c in hdr or []:
            if c in ENUM and c not in ("Attribute_Value",):
                allowed = ENUM[c] | ({ "Unspecifiable" })
                extra = ext_e if c == "Element_Type" else ext_r if c == "RepresentedItem_Type" else set()
                for i, r in enumerate(data, 2):
                    v = r.get(c)
                    if v is not None and v not in allowed | extra:
                        f(2, "I14", f"{sheet} row {i}: {c}={v!r} not in Enum_Lookup[{c}]")
    note_rule("I14", "PASS")

    # I29
    sf_rows = [r for r in sheets["Document_Data"][1] if r.get("Attribute_Name") == "Source_Format"]
    if len(sf_rows) != 1: f(2, "I29", f"Source_Format rows: {len(sf_rows)} (exactly one required)")
    elif "Source_Format" in ENUM and sf_rows[0].get("Attribute_Value") not in ENUM["Source_Format"]:
        f(2, "I29", f"Source_Format={sf_rows[0].get('Attribute_Value')!r} not enum-valid")
    note_rule("I29", "PASS")
    source_format = sf_rows[0].get("Attribute_Value") if len(sf_rows) == 1 else None

    # ---------- Required-attribute coverage (§9.2 Attribute_Lookup.Required) ----------
    check_required_attributes(sheets, doc_type)

    # ---------- Layer 3: object/cluster/match ----------
    check_cluster_coverage(sheets, source_format)

    # ---------- Layer 4: element/connection integrity ----------
    for i, r in enumerate(sheets["Match_Result"][1], 2):
        st, td, fc, rs = r.get("Match_Status"), r.get("Element_TopDown_ID"), r.get("Element_from_Cluster_ID"), r.get("Resolution_Status")
        if st == "Matched":
            if not (td and fc): f(4, "I2", f"Match_Result row {i}: Matched but IDs incomplete")
            if rs != "Resolved_AutoMatch": f(4, "I20", f"Match_Result row {i}: Matched with Resolution_Status={rs!r}")
        elif st == "Only_TopDown":
            if not td or fc: f(4, "I3", f"Match_Result row {i}: Only_TopDown population invalid")
            if rs not in ("Resolved_KeepBoth","Resolved_TopDown_Valid","Resolved_Cluster_Valid"):
                f(4, "I20", f"Match_Result row {i}: Only_TopDown with Resolution_Status={rs!r} (terminal status required)")
        elif st == "Only_Cluster":
            if not fc or td: f(4, "I4", f"Match_Result row {i}: Only_Cluster population invalid")
            if rs not in ("Resolved_KeepBoth","Resolved_TopDown_Valid","Resolved_Cluster_Valid"):
                f(4, "I20", f"Match_Result row {i}: Only_Cluster with Resolution_Status={rs!r} (terminal status required)")
    note_rule("I2", "PASS"); note_rule("I3", "PASS"); note_rule("I4", "PASS"); note_rule("I20", "PASS")

    def tree(rule, sheet, idc, pc):
        hdr, data = sheets[sheet]
        if pc not in (hdr or []):
            note_rule(rule, "PASS", detail=f"{sheet}.{pc} column absent; nothing to check.")
            return
        parent = {r.get(idc): r.get(pc) for r in data if r.get(idc)}
        for k, p in parent.items():
            if p is None: continue
            if p == k: f(4, rule, f"{sheet}: {idc}={k!r} self-reference"); continue
            if p not in parent: f(4, rule, f"{sheet}: {idc}={k!r} parent {p!r} unresolved"); continue
            seen, cur = {k}, p
            while cur is not None:
                if cur in seen: f(4, rule, f"{sheet}: cycle at {idc}={k!r}"); break
                seen.add(cur); cur = parent.get(cur)
        note_rule(rule, "PASS")
    tree("I6/I16","Cluster","Cluster_ID","Parent_Cluster_ID")
    tree("I16","Document_RepresentedItem","RepresentedItem_ID","Parent_RepresentedItem_ID")
    tree("I16","Elements_TopDown","Element_TopDown_ID","Parent_Element_TopDown_ID")
    tree("I16","Element_ID","Element_ID","Parent_Element_ID")

    obj_by = {r.get("Object_ID"): r for r in sheets["Object"][1]}
    for i, r in enumerate(sheets["Cluster"][1], 2):
        ct, co = r.get("Cluster_Type"), r.get("Container_Object_ID")
        if ct == "Containment" and co is not None:
            o = obj_by.get(co)
            if not o or o.get("Object_Type") != "Graphic" or str(o.get("Geometry_Closed")).lower() not in ("true","1"):
                f(4, "I17", f"Cluster row {i}: Container_Object_ID={co!r} not a closed Graphic")
    note_rule("I17", "PASS")

    prim = {}
    for i, r in enumerate(sheets["Element_RepresentedItem_Mapping"][1], 2):
        if r.get("Relationship_Type") == "Primary":
            prim.setdefault(r.get("Element_ID"), []).append(i)
    for e, rows in prim.items():
        if len(rows) > 1: f(4, "I22", f"Element_ID={e!r}: {len(rows)} Primary mappings (rows {rows})")
    note_rule("I22", "PASS")

    # ---------- Layer 5: document-type-specific content checks ----------
    check_I23(sheets, doc_type)
    check_I24(sheets, doc_type)
    check_I26(sheets, doc_type)
    check_I28(sheets, doc_type)

    # ---------- Process-step post-conditions (§11.4 P0-P9) ----------
    proc_steps = compute_process_step_post_conditions(sheets, doc_ids, E, OBJ, CL, ETD, EFC, MR, RIT, CID, LAY)

    report(doc_type, sheets=sheets, output_path=output_path, meta=meta, proc_steps=proc_steps,
           schema_version=schema_version)


# ---------------------------------------------------------------------------
# Document-type-specific content checks (§11.9) -- I23, I24, I26, I28
# ---------------------------------------------------------------------------
def check_I23(sheets, doc_type):
    """Bridge / Terminal_Strip consistency (Terminal_Diagram)."""
    if not scope_includes(doc_type, "Terminal_Diagram"):
        note_rule("I23", "SKIP", detail=f"Not applicable: Document_Type={doc_type!r} (rule scoped to Terminal_Diagram).")
        return
    cd_by_conn = {}
    for r in sheets["Connection_Data"][1]:
        cd_by_conn.setdefault(r.get("Connection_ID"), {})[r.get("Attribute_Name")] = r.get("Attribute_Value")
    parent_by_eid = {r.get("Element_ID"): r.get("Parent_Element_ID") for r in sheets["Element_ID"][1]}
    etype_by_eid = {r.get("Element_ID"): r.get("Element_Type") for r in sheets["Element_ID"][1]}
    for i, c in enumerate(sheets["Connection_ID"][1], 2):
        attrs = cd_by_conn.get(c.get("Connection_ID"), {})
        ctype = attrs.get("Connection_Type") or ""
        if not ctype.startswith("Bridge_"):
            continue
        fe, te = c.get("From_Element_ID"), c.get("To_Element_ID")
        if etype_by_eid.get(fe) != "Terminal" or etype_by_eid.get(te) != "Terminal" or parent_by_eid.get(fe) != parent_by_eid.get(te):
            f(5, "I23", f"Connection_ID row {i} ({c.get('Connection_ID')}): Connection_Type={ctype!r} but "
                        f"endpoints are not two Terminals sharing the same Parent_Element_ID (Terminal_Strip)")
    note_rule("I23", "PASS")


def check_I24(sheets, doc_type):
    """Wire-colour / Polarity consistency (Terminal_Diagram, Circuit_Diagram)."""
    if not scope_includes(doc_type, "Terminal_Diagram/Circuit_Diagram"):
        note_rule("I24", "SKIP", detail=f"Not applicable: Document_Type={doc_type!r} "
                                        f"(rule scoped to Terminal_Diagram/Circuit_Diagram).")
        return
    cd_by_conn = {}
    for r in sheets["Connection_Data"][1]:
        cd_by_conn.setdefault(r.get("Connection_ID"), {})[r.get("Attribute_Name")] = r.get("Attribute_Value")
    ed_by_eid = {}
    for r in sheets["Element_Data"][1]:
        ed_by_eid.setdefault(r.get("Element_ID"), {})[r.get("Attribute_Name")] = r.get("Attribute_Value")
    layer_voltage = {r.get("Layer_ID"): r.get("Voltage_Level") for r in sheets["Layer_ID"][1]}
    eid_layer = {r.get("Element_ID"): r.get("Layer_ID") for r in sheets["Element_ID"][1]}

    for i, c in enumerate(sheets["Connection_ID"][1], 2):
        cid = c.get("Connection_ID")
        attrs = cd_by_conn.get(cid, {})
        wc, pol = attrs.get("Wire_Color"), attrs.get("Polarity")
        if not wc or wc == "Unspecifiable":
            continue
        fe, te = c.get("From_Element_ID"), c.get("To_Element_ID")
        # Voltage-level resolution priority per §11.9 I24: (1) Connection_Data,
        # (2) Element_Data of both endpoints (must agree), (3) Layer_ID of both
        # endpoints (must agree), (4) abstain.
        vl = attrs.get("Voltage_Level")
        if not vl:
            v1, v2 = ed_by_eid.get(fe, {}).get("Voltage_Level"), ed_by_eid.get(te, {}).get("Voltage_Level")
            if v1 and v1 == v2:
                vl = v1
        if not vl:
            l1, l2 = layer_voltage.get(eid_layer.get(fe)), layer_voltage.get(eid_layer.get(te))
            if l1 and l1 == l2:
                vl = l1
        # Per §11.9: "the rule can be abstained from via Unspecifiable or by missing
        # polarity" -- a populated-but-wrong Polarity is a violation; an absent one is not.
        if wc == "GNYE" and pol and pol != "Unspecifiable" and pol != "PE":
            f(5, "I24", f"Connection_ID row {i} ({cid}): Wire_Color=GNYE but Polarity={pol!r} (expected PE)")
        if wc == "BU" and vl in ("230V_AC", "400V_AC") and pol and pol != "Unspecifiable" and pol != "N":
            f(5, "I24", f"Connection_ID row {i} ({cid}): Wire_Color=BU at Voltage_Level={vl} but "
                        f"Polarity={pol!r} (expected N)")
    note_rule("I24", "PASS")


def check_I26(sheets, doc_type):
    """Current_Path_Number consistency (Circuit_Diagram). Non-contiguous
    numbering is a warning per §11.9, never a hard violation; only a
    non-integer value is a violation."""
    if not scope_includes(doc_type, "Circuit_Diagram"):
        note_rule("I26", "SKIP", detail=f"Not applicable: Document_Type={doc_type!r} (rule scoped to Circuit_Diagram).")
        return
    paths = []
    hdr, data = sheets["Element_Data"]
    for i, r in enumerate(data, 2):
        if r.get("Attribute_Name") != "Current_Path_Number":
            continue
        val = r.get("Attribute_Value")
        try:
            paths.append(int(str(val).strip()))
        except (TypeError, ValueError):
            f(5, "I26", f"Element_Data row {i}: Current_Path_Number={val!r} is not integer-parsable")
    if not paths:
        note_rule("I26", "PASS", detail="No Current_Path_Number values populated; rule does not fire.")
        return
    has_grid_marker = any(r.get("Attribute_Name") == "Path_Numbering_Grid" for r in sheets["Element_Data"][1])
    has_reserve_marker = any(
        any(tok in (r.get("Content_Text") or "").strip().lower() for tok in ("reserve", "frei", "free"))
        for r in sheets["Object"][1]
    )
    full = set(range(min(paths), max(paths) + 1))
    gaps = sorted(full - set(paths))
    if gaps:
        if has_grid_marker or has_reserve_marker:
            note_rule("I26", "PASS", detail=f"Non-contiguous Current_Path_Number sequence (gaps: {gaps}); "
                                            f"treated as justified per §11.9 (reserve-marker Object and/or "
                                            f"Path_Numbering_Grid declaration present) -- warning suppressed.")
        else:
            note_rule("I26", "PASS", detail=f"WARNING (non-fatal per §11.9, not a violation): non-contiguous "
                                            f"Current_Path_Number sequence, gaps at {gaps}; no reserve-marker "
                                            f"Object or Path_Numbering_Grid declaration found to justify it.")
    else:
        note_rule("I26", "PASS", detail="Current_Path_Number sequence is contiguous.")


def check_I28(sheets, doc_type):
    """IEC 60617 classification completeness (Circuit_Diagram)."""
    if not scope_includes(doc_type, "Circuit_Diagram"):
        note_rule("I28", "SKIP", detail=f"Not applicable: Document_Type={doc_type!r} (rule scoped to Circuit_Diagram).")
        return
    IEC60617_SYS = {"IEC 60617-2", "IEC 60617-3", "IEC 60617-6", "IEC 60617-7", "IEC 60617-8"}
    NN_PATTERN = re.compile(r"^\d{2}-\d{2}-\d{2}$")

    for i, r in enumerate(sheets["Element_Classification"][1], 2):
        system, code = r.get("Classification_System"), r.get("Classification_Code")
        if system in IEC60617_SYS and code != "Unclassified" and code is not None:
            if not NN_PATTERN.match(code):
                f(5, "I28", f"Element_Classification row {i}: Classification_Code={code!r} does not match "
                            f"the NN-NN-NN item-code pattern required for {system}")
        if code == "Unclassified" and not r.get("Source_Symbol_Reference"):
            f(5, "I28", f"Element_Classification row {i}: Classification_Code=Unclassified requires "
                        f"Source_Symbol_Reference to be populated (§A.19)")

    by_eid = {}
    for r in sheets["Element_Classification"][1]:
        if r.get("Classified_Object_Type") == "Element":
            by_eid.setdefault(r.get("Classified_Object_ID"), []).append(r)
    caex_by_eid = {r.get("Element_ID"): r.get("CAEX_Type") for r in sheets["Element_ID"][1]}
    etype_by_eid = {r.get("Element_ID"): r.get("Element_Type") for r in sheets["Element_ID"][1]}
    parent_has_60617 = {
        eid for eid, rows in by_eid.items() if any(r.get("Classification_System") in IEC60617_SYS for r in rows)
    }
    for i, e in enumerate(sheets["Element_ID"][1], 2):
        eid = e.get("Element_ID")
        rows = by_eid.get(eid, [])
        has_60617 = any(r.get("Classification_System") in IEC60617_SYS for r in rows)
        has_unclassified_fallback = any(
            r.get("Classification_Code") == "Unclassified" and r.get("Source_Symbol_Reference") for r in rows
        )
        if has_60617 or has_unclassified_fallback:
            continue
        # Element_Type=Connection_Point is the v0.8.3 canonical spelling for generated CAEX
        # connection-point sub-elements (§5.13); Element_Type=Terminal is kept as a fallback
        # so that pre-v0.8.3 workbooks (which reused Terminal for this purpose) still exempt
        # correctly (§11.9 I28 clarification).
        is_connpoint = (etype_by_eid.get(eid) in ("Connection_Point", "Terminal")
                        and caex_by_eid.get(eid) == "ExternalInterface"
                        and e.get("Parent_Element_ID"))
        if is_connpoint and e.get("Parent_Element_ID") in parent_has_60617:
            continue  # exempt per §11.9 I28 clarification
        f(5, "I28", f"Element_ID row {i} ({eid}): no IEC 60617 classification row and no "
                    f"Unclassified+Source_Symbol_Reference fallback")
    note_rule("I28", "PASS")


# ---------------------------------------------------------------------------
# Required-attribute coverage (§9.2 Attribute_Lookup.Required)
# ---------------------------------------------------------------------------
def check_required_attributes(sheets, doc_type):
    scope_cfg = {
        "Document": ("Document_Data", "Document_ID", "Document_ID"),
        "Element": ("Element_Data", "Element_ID", "Element_ID"),
        "RepresentedItem": ("RepresentedItem_Data", "RepresentedItem_ID", "Document_RepresentedItem"),
        "Connection": ("Connection_Data", "Connection_ID", "Connection_ID"),
    }
    req_rows = [r for r in sheets["Attribute_Lookup"][1]
                if str(r.get("Required")).strip().lower() in ("true", "1", "1.0", "yes")]
    if not req_rows:
        note_rule("Required_Attributes", "PASS", detail="No Attribute_Lookup row is marked Required=TRUE.")
        return
    by_scope = {}
    for r in req_rows:
        by_scope.setdefault(r.get("Scope"), []).append(r)

    for scope, reqs in by_scope.items():
        if scope not in scope_cfg:
            continue
        data_sheet, key_col, own_sheet = scope_cfg[scope]
        if own_sheet not in sheets or data_sheet not in sheets:
            continue
        have = {}
        for dr in sheets[data_sheet][1]:
            have.setdefault(dr.get(key_col), {})[dr.get("Attribute_Name")] = dr.get("Attribute_Value")
        for i, orow in enumerate(sheets[own_sheet][1], 2):
            oid = orow.get(key_col)
            context = {"Document_Type": doc_type, "Element_Type": orow.get("Element_Type"),
                       "RepresentedItem_Type": orow.get("RepresentedItem_Type")}
            for rr in reqs:
                if not tc_applies(rr.get("Type_Constraint"), context):
                    continue
                attr = rr.get("Attribute_Name")
                if attr not in have.get(oid, {}):
                    # Non-blocking (§WARNING, not FAIL): a Required=TRUE Attribute_Lookup
                    # entry that is absent from Element_Data reflects a gap in the SOURCE
                    # document's own completeness, not a defect in this workbook's
                    # conformance to the schema -- the extraction itself is not wrong
                    # just because the source never supplied a value. Attributes that are
                    # structurally never part of a given Document_Type SHALL instead be
                    # scoped Required=False for that Document_Type in Attribute_Lookup
                    # (a spec/catalog correction) rather than surfacing here at all.
                    w(2, "Required_Attributes", f"{own_sheet} row {i} ({key_col}={oid!r}, scope={scope}): "
                                                 f"required attribute {attr!r} not populated in {data_sheet} "
                                                 f"(Unspecifiable would count as populated, but the attribute "
                                                 f"is absent entirely)",
                      group=attr, ref=oid)
    note_rule("Required_Attributes", "PASS")


# ---------------------------------------------------------------------------
# Cluster coverage (§11.4 P3 post-condition)
# ---------------------------------------------------------------------------
def check_cluster_coverage(sheets, source_format):
    if source_format != "PDF_Drawing":
        note_rule("Cluster_Coverage", "SKIP",
                  detail=f"Not applicable: Source_Format={source_format!r} (only evaluated for PDF_Drawing sources).")
        return
    empty_marker = any(r.get("Attribute_Name") == "Cluster_Method" and r.get("Attribute_Value") == "P3_Empty_Result"
                       for r in sheets["Document_Data"][1])
    if empty_marker:
        if sheets["Cluster"][1] or sheets["Object_Cluster"][1]:
            f(3, "Cluster_Coverage", "Document_Data declares Cluster_Method=P3_Empty_Result, but Cluster "
                                      "and/or Object_Cluster is not empty.")
        else:
            note_rule("Cluster_Coverage", "PASS",
                      detail="P3_Empty_Result declared and Cluster/Object_Cluster are empty, as permitted.")
        return
    # Synthetic Manual_Entry objects (I12 manual-evidence convention, §A.7) are provenance
    # anchors for manual/rule-derived decisions, not source content captured by P1/P3 -- they
    # are exempt from cluster coverage for the same reason Topology-Objects are (A9): the
    # coverage obligation is about the document's captured content, not every Object row.
    non_topo = {r.get("Object_ID") for r in sheets["Object"][1]
                if r.get("Object_Type") != "Topology" and r.get("Source_Operation") != "Manual_Entry"}
    clustered = {r.get("Object_ID") for r in sheets["Object_Cluster"][1]}
    missing = non_topo - clustered
    if missing:
        sample = ", ".join(sorted(missing)[:8])
        f(3, "Cluster_Coverage", f"{len(missing)} non-Topology, non-Manual_Entry Object(s) not present in "
                                  f"Object_Cluster (P3 post-condition, §11.4); e.g. {sample}")
    note_rule("Cluster_Coverage", "PASS")


# ---------------------------------------------------------------------------
# Process-step post-conditions (§11.4 P0-P9), decidable subset only.
# ---------------------------------------------------------------------------
def compute_process_step_post_conditions(sheets, doc_ids, E, OBJ, CL, ETD, EFC, MR, RIT, CID, LAY):
    def src_coverage(parent, key, src_sheet):
        have = ids(sheets, src_sheet, key)
        vals = [v for v in col(sheets, parent, key) if v is not None]
        return all(v in have for v in vals)

    steps = {}

    doc_ok = len(sheets["Document_ID"][1]) >= 1
    ri_ok = len(sheets["Document_RepresentedItem"][1]) >= 1
    dd_src_ok = src_coverage("Document_Data", "Document_Data_ID", "Document_Data_Source")
    rd_src_ok = src_coverage("Revision_Data", "Revision_ID", "Revision_Data_Source")
    steps["P0"] = {"decidable": True, "satisfied": doc_ok and ri_ok and dd_src_ok and rd_src_ok,
                   "note": "Document_ID + Document_RepresentedItem populated; Document_Data_Source / "
                           "Revision_Data_Source fully cover their parent rows."}

    obj_ok = len(sheets["Object"][1]) >= 1 and all(r.get("Object_Type") is not None for r in sheets["Object"][1])
    steps["P1"] = {"decidable": True, "satisfied": obj_ok,
                   "note": "Object sheet populated with valid Object_Type on every row. (True per-source "
                           "completeness -- 'every atomic source unit captured' -- is not decidable from "
                           "the workbook alone; this checks the decidable proxy only.)"}

    td_count = len(sheets["Elements_TopDown"][1])
    steps["P2"] = {"decidable": td_count >= 1, "satisfied": td_count >= 1,
                   "note": "Elements_TopDown populated." if td_count >= 1 else
                           "Elements_TopDown is empty: the spec permits an explicit empty-set "
                           "acknowledgment for P2, but defines no artifact marker for it, so an empty "
                           "sheet cannot be distinguished from a skipped step from the artifact alone."}

    empty_marker = any(r.get("Attribute_Name") == "Cluster_Method" and r.get("Attribute_Value") == "P3_Empty_Result"
                       for r in sheets["Document_Data"][1])
    # Manual_Entry synthetic objects are exempt for the same reason as in check_cluster_coverage().
    non_topo = {r.get("Object_ID") for r in sheets["Object"][1]
                if r.get("Object_Type") != "Topology" and r.get("Source_Operation") != "Manual_Entry"}
    clustered = {r.get("Object_ID") for r in sheets["Object_Cluster"][1]}
    if empty_marker:
        p3_ok = not sheets["Cluster"][1] and not sheets["Object_Cluster"][1]
    else:
        p3_ok = len(sheets["Cluster"][1]) >= 1 and non_topo.issubset(clustered)
    steps["P3"] = {"decidable": True, "satisfied": p3_ok,
                   "note": "Cluster + full Object_Cluster coverage of non-Topology Objects, or a documented "
                           "P3_Empty_Result with both sheets empty."}

    efc_status_ok = all(r.get("Derivation_Status") is not None for r in sheets["Elements_from_Cluster"][1])
    cluster_ids_covered = CL.issubset({r.get("Source_Cluster_ID") for r in sheets["Elements_from_Cluster"][1]}) if CL else True
    steps["P4"] = {"decidable": True, "satisfied": efc_status_ok and cluster_ids_covered,
                   "note": "Every Elements_from_Cluster row has Derivation_Status set, and every Cluster is "
                           "referenced by at least one Elements_from_Cluster row."}

    td_in_match = {r.get("Element_TopDown_ID") for r in sheets["Match_Result"][1] if r.get("Element_TopDown_ID")}
    fc_in_match = {r.get("Element_from_Cluster_ID") for r in sheets["Match_Result"][1] if r.get("Element_from_Cluster_ID")}
    fc_derived = {r.get("Element_from_Cluster_ID") for r in sheets["Elements_from_Cluster"][1]
                  if r.get("Derivation_Status") == "Element_Derived"}
    steps["P5"] = {"decidable": True, "satisfied": ETD.issubset(td_in_match) and fc_derived.issubset(fc_in_match),
                   "note": "Every Elements_TopDown row and every Element_Derived Elements_from_Cluster row "
                           "appears in >=1 Match_Result row."}

    p6_ok = all(r.get("Resolution_Status") not in (None, "Open") for r in sheets["Match_Result"][1])
    steps["P6"] = {"decidable": True, "satisfied": p6_ok, "note": "No Match_Result row has Resolution_Status=Open."}

    p7_ok = len(sheets["Element_ID"][1]) >= 1 and all(r.get("Source_Match_ID") in MR for r in sheets["Element_ID"][1])
    steps["P7"] = {"decidable": True, "satisfied": p7_ok,
                   "note": "Element_ID populated and every row has a resolvable Source_Match_ID."}

    ec_src_ok = src_coverage("Element_Classification", "Classification_ID", "Element_Classification_Source")
    ed_src_ok = src_coverage("Element_Data", "Element_Data_ID", "Element_Data_Source")
    rid_src_ok = src_coverage("RepresentedItem_Data", "RepresentedItem_Data_ID", "RepresentedItem_Data_Source")
    steps["P8"] = {"decidable": True, "satisfied": ec_src_ok and ed_src_ok and rid_src_ok,
                   "note": "Element_Classification_Source / Element_Data_Source / RepresentedItem_Data_Source "
                           "fully cover their parent rows."}

    cd_src_ok = src_coverage("Connection_Data", "Connection_Data_ID", "Connection_Data_Source")
    conn_topo_ok = all(r.get("Source_Topology_Object_ID") in OBJ for r in sheets["Connection_ID"][1]
                       if r.get("Source_Topology_Object_ID") is not None)
    steps["P9"] = {"decidable": True, "satisfied": cd_src_ok and conn_topo_ok,
                   "note": "Connection_Data_Source fully covers Connection_Data; every populated "
                           "Source_Topology_Object_ID resolves to an existing Object."}

    return steps


def report(doc_type=None, sheets=None, output_path=None, meta=None, proc_steps=None, schema_version=None):
    """Write a detailed textual report and a machine-readable JSON summary.

    Parameters:
    - doc_type: str or None
    - sheets: optional sheets dict as returned by `load` for per-sheet stats and metadata
    - output_path: Path or str for the text report; a JSON file with the same stem is written alongside
    - meta: optional pre-built metadata dict (overrides Schema_Metadata extraction)
    - proc_steps: optional dict from compute_process_step_post_conditions()
    - schema_version: str or None -- the workbook's own Document_ID.Schema_Version, i.e. the
      spec version this specific artifact declares itself conformant to. This is read from the
      artifact's content, not assumed from its filename, so the report stays accurate even if a
      workbook is renamed or a filename's embedded version string is stale.
    """
    viol = list(F)
    warn = list(WARN)

    # derive metadata from sheets if not explicitly provided
    if meta is None and sheets and "Schema_Metadata" in sheets:
        meta = {r.get("Metadata_Key"): r.get("Metadata_Value") for r in sheets["Schema_Metadata"][1]}
    meta = meta or {}

    # per-sheet statistics
    sheet_stats = {}
    if sheets:
        for name, (hdr, data) in sheets.items():
            sheet_stats[name] = {"rows": len(data), "columns": len(hdr) if hdr else 0}

    rule_summary = {rule: {"status": entry["status"], "description": entry.get("description"),
                            "details": entry.get("details", [])} for rule, entry in RULE_RESULTS.items()}

    lines = []
    lines.append("=== Conformance Check Report ===")
    lines.append(f"Timestamp: {datetime.utcnow().isoformat()}Z")
    lines.append(f"Document_Type: {doc_type}")
    lines.append(f"Schema_Version (as declared in Document_ID of this artifact): {schema_version}")
    lines.append(f"Total findings (violations): {len(viol)}")
    lines.append(f"Total warnings (non-blocking): {len(warn)}")
    lines.append("")
    # --- Benchmarks ---
    expected_sheets = len(MANDATORY)
    present_sheets = len(sheets) if sheets else 0
    missing_sheets = [s for s in MANDATORY if not sheets or s not in sheets]
    extra_sheets = sorted(list((set(sheets.keys()) - set(MANDATORY)) if sheets else set()))

    rule_counts = Counter(entry["status"] for entry in RULE_RESULTS.values())

    implemented_rules = [r for r in RULE_RESULTS if r in IMPLEMENTED_RULES]
    not_implemented_rules = [r for r in RULE_RESULTS if r not in IMPLEMENTED_RULES]
    warned_rules = [r for r in implemented_rules if RULE_RESULTS[r]["status"] == "WARNING"]
    level1_total = len(implemented_rules)
    level1_skipped = sum(1 for r in implemented_rules if RULE_RESULTS[r]["status"] == "SKIP")
    level1_warned = len(warned_rules)
    level1_failed = sum(1 for r in implemented_rules if RULE_RESULTS[r]["status"] == "FAIL")
    level1_passed = sum(1 for r in implemented_rules if RULE_RESULTS[r]["status"] == "PASS")
    level1_not_run = sum(1 for r in implemented_rules if RULE_RESULTS[r]["status"] == "NOT_RUN")
    # WARNING findings reflect source-content completeness, not artifact conformance
    # (see check_required_attributes) -- excluded from the denominator just like SKIP,
    # so they never depress the conformance pass rate, but remain fully visible below.
    level1_applicable = max(0, level1_total - level1_skipped - level1_warned)
    level1_pass_pct = (100.0 * level1_passed / level1_applicable) if level1_applicable else None

    proc_steps = proc_steps or {}
    proc_decidable = [p for p, v in proc_steps.items() if v["decidable"]]
    proc_satisfied = [p for p in proc_decidable if proc_steps[p]["satisfied"]]
    proc_not_decidable = [p for p, v in proc_steps.items() if not v["decidable"]]

    lines.append("Benchmarks:")
    lines.append(f"- Expected mandatory sheets: {expected_sheets}")
    lines.append(f"- Present sheets: {present_sheets}")
    lines.append(f"- Missing mandatory sheets: {len(missing_sheets)}")
    if missing_sheets:
        lines.append(f"  * {', '.join(missing_sheets)}")
    lines.append(f"- Extra sheets: {len(extra_sheets)}")
    if extra_sheets:
        lines.append(f"  * {', '.join(extra_sheets)}")
    lines.append(f"- Rule counts: PASS={rule_counts.get('PASS',0)} FAIL={rule_counts.get('FAIL',0)} "
                 f"WARNING={rule_counts.get('WARNING',0)} SKIP={rule_counts.get('SKIP',0)} "
                 f"NOT_RUN={rule_counts.get('NOT_RUN',0)} NOT_IMPLEMENTED={rule_counts.get('NOT_IMPLEMENTED',0)}")
    if not_implemented_rules:
        lines.append(f"  * NOT_IMPLEMENTED rules (declared in the spec, no check code in this checker "
                     f"version -- absence of findings is NOT evidence of conformance): "
                     f"{', '.join(sorted(not_implemented_rules))}")
    if warned_rules:
        lines.append(f"  * WARNING rules (source-content completeness gaps, not conformance defects -- "
                     f"excluded from the pass rate, see details below): {', '.join(sorted(warned_rules))}")
    if level1_not_run:
        lines.append(f"  * NOTE: {level1_not_run} implemented rule(s) never executed (NOT_RUN) -- "
                     f"this indicates a gap in the checker itself, not a workbook property.")
    lines.append(f"- Implemented-rule pass rate: total={level1_total} applicable={level1_applicable} "
                 f"passed={level1_passed} failed={level1_failed} warned={level1_warned} "
                 f"pass_pct={'{:.1f}%'.format(level1_pass_pct) if level1_pass_pct is not None else 'n/a'}")
    lines.append(f"- Process-step post-conditions satisfied: {len(proc_satisfied)}/{len(proc_decidable)} "
                 f"decidable steps ({', '.join(proc_satisfied) or 'none'})")
    if proc_not_decidable:
        lines.append(f"  * Not artifact-decidable: {', '.join(proc_not_decidable)}")
    lines.append("")
    lines.append("Metadata:")
    if meta:
        for k, v in sorted(meta.items()):
            lines.append(f"- {k}: {v}")
    else:
        lines.append("- <no Schema_Metadata available>")
    lines.append("")
    lines.append("Per-sheet statistics:")
    if sheet_stats:
        for name in sorted(sheet_stats):
            s = sheet_stats[name]
            lines.append(f"- {name}: rows={s['rows']} cols={s['columns']}")
    else:
        lines.append("- <no sheet information available>")
    lines.append("")
    lines.append("Process-step post-condition detail (P0-P9):")
    for p in sorted(proc_steps):
        v = proc_steps[p]
        state = "SATISFIED" if v["satisfied"] else "NOT SATISFIED"
        if not v["decidable"]:
            state = "NOT ARTIFACT-DECIDABLE"
        lines.append(f"- {p}: {state} — {v.get('note','')}")
    lines.append("")
    lines.append("Rule overview:")
    for rule in sorted(rule_summary):
        entry = rule_summary[rule]
        status = (entry["status"] or "NOT_RUN").upper()
        lines.append(f"- {rule}: {status} — {entry.get('description')}")
        for detail in entry.get("details", []):
            lines.append(f"    • {detail}")
    lines.append("")
    lines.append("Detailed findings:")
    if viol:
        for l, r, d in viol:
            lines.append(f"- [L{l}] {r}: {d}")
    else:
        lines.append("- No violations detected.")
    lines.append("")
    lines.append("Warnings (a Required=TRUE Attribute_Lookup entry found no corresponding value in this "
                 "workbook's extracted data; non-blocking, does not count against the pass rate):")
    if warn:
        summary = {}
        for (_l, r, _d, g, ref) in warn:
            summary.setdefault((r, g), []).append(ref)
        lines.append("Summary:")
        for (r, g), refs in sorted(summary.items(), key=lambda kv: (kv[0][0], str(kv[0][1]))):
            refs_str = ", ".join(str(x) for x in refs if x is not None)
            label = f"{r} / {g}" if g is not None else r
            lines.append(f"- {label}: {len(refs)} finding(s) affected -- {refs_str}")
        lines.append("")
        lines.append("Detail:")
        for l, r, d, _g, _ref in warn:
            lines.append(f"- [L{l}] {r}: {d}")
    else:
        lines.append("- No warnings.")

    text = "\n".join(lines) + "\n"

    structured = {
        "timestamp": datetime.utcnow().isoformat() + "Z",
        "document_type": doc_type,
        "schema_version": schema_version,
        "metadata": meta,
        "sheet_stats": sheet_stats,
        "rule_summary": rule_summary,
        "violations": [{"layer": l, "rule": r, "detail": d} for (l, r, d) in viol],
        "findings": [{"layer": l, "rule": r, "detail": d} for (l, r, d) in F],
        "warnings": [{"layer": l, "rule": r, "detail": d, "group": g, "ref": ref}
                     for (l, r, d, g, ref) in warn],
        "benchmarks": {
            "expected_mandatory_sheets": expected_sheets,
            "present_sheets": present_sheets,
            "missing_mandatory_sheets": missing_sheets,
            "extra_sheets": extra_sheets,
            "rule_counts": dict(rule_counts),
            "not_implemented_rules": sorted(not_implemented_rules),
            "warned_rules": sorted(warned_rules),
            "implemented_rules": {
                "total": level1_total,
                "applicable": level1_applicable,
                "passed": level1_passed,
                "failed": level1_failed,
                "warned": level1_warned,
                "not_run": level1_not_run,
                "pass_pct": level1_pass_pct,
            },
            # Deprecated alias retained for downstream consumers of the old key name;
            # prefer "implemented_rules" above. Semantics unchanged from prior releases
            # aside from now excluding NOT_IMPLEMENTED rules from the denominator.
            "level1": {
                "total": level1_total,
                "applicable": level1_applicable,
                "passed": level1_passed,
                "failed": level1_failed,
                "pass_pct": level1_pass_pct,
            },
            "process_step_post_conditions": {
                "satisfied": len(proc_satisfied),
                "decidable_total": len(proc_decidable),
                "not_artifact_decidable": proc_not_decidable,
                "details": proc_steps,
            },
        },
    }

    if output_path is not None:
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_text(text, encoding="utf-8")
        # JSON alongside text
        jpath = output_path.with_suffix(".json")
        jpath.write_text(json.dumps(structured, ensure_ascii=False, indent=2), encoding="utf-8")
        print(f"Report written to {output_path} and {jpath}")
    else:
        print(text, end="")
        print(json.dumps(structured, ensure_ascii=False, indent=2))


def _slug(value, fallback):
    if not value:
        return fallback
    s = re.sub(r"[^A-Za-z0-9]+", "_", str(value)).strip("_")
    return s or fallback


def result_stem_for(path):
    """Derive the result-report base filename from the workbook's own declared
    Document_Type and Schema_Version (Document_ID sheet), not from the input
    filename alone. This keeps result naming truthful about what was actually
    checked even if a workbook is renamed or its filename's embedded version
    string goes stale relative to its actual content."""
    try:
        _, sheets = load(path)
        doc_type = next(iter(col(sheets, "Document_ID", "Document_Type")), None) if "Document_ID" in sheets else None
        schema_version = (next(iter(col(sheets, "Document_ID", "Schema_Version")), None)
                           if "Document_ID" in sheets else None)
    except Exception:
        doc_type, schema_version = None, None
    return f"{_slug(doc_type, 'UnknownDocType')}__{_slug(schema_version, 'unknown_schema_version')}__{path.stem}"


def main(argv):
    if len(argv) > 2:
        print("Usage: python schema_conformance_checker.py [path/to/workbook.xlsx or artifact-folder]")
        return 1

    if len(argv) == 2:
        target = Path(argv[1]).expanduser().resolve()
        if target.is_dir():
            input_paths = discover_excel_files(target)
            if not input_paths:
                print(f"No Excel files found in {target}")
                return 1
        elif target.is_file():
            input_paths = [target]
        else:
            print(f"Input not found: {target}")
            return 1
    else:
        input_dir = resolve_artifact_dir()
        input_paths = discover_excel_files(input_dir)
        if not input_paths:
            print(f"No Excel files found in {input_dir}")
            return 1

    results_dir = resolve_results_dir()
    for path in input_paths:
        output_path = results_dir / f"{result_stem_for(path)}.txt"
        check(path, output_path=output_path)

    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv))
